"""
Export utilities for generating PDF and Excel reports.
"""
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from django.conf import settings
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak


def export_letters_to_excel(letters, title="Letters Report"):
    """
    Export letters to Excel format using openpyxl.
    """
    output = BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Letters"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    # Headers
    headers = [
        "Reference No", "Direction", "Date", "Subject", "Sender/Recipient",
        "Category", "Priority", "Status", "Department", "Assigned To", "Due Date"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Data rows
    for row_num, letter in enumerate(letters, 2):
        worksheet.cell(row=row_num, column=1, value=letter.reference_no or "").border = border
        worksheet.cell(row=row_num, column=2, value=letter.get_direction_display()).border = border
        worksheet.cell(row=row_num, column=3, value=letter.date.strftime("%Y-%m-%d") if letter.date else "").border = border
        worksheet.cell(row=row_num, column=4, value=letter.subject).border = border
        
        sender_recipient = letter.sender if letter.direction == 'INCOMING' else letter.recipient
        worksheet.cell(row=row_num, column=5, value=sender_recipient or "").border = border
        
        worksheet.cell(row=row_num, column=6, value=letter.category.name if letter.category else "").border = border
        worksheet.cell(row=row_num, column=7, value=letter.get_priority_display()).border = border
        worksheet.cell(row=row_num, column=8, value=letter.get_status_display()).border = border
        worksheet.cell(row=row_num, column=9, value=letter.assigned_department.name if letter.assigned_department else "").border = border
        worksheet.cell(row=row_num, column=10, value=str(letter.assigned_person) if letter.assigned_person else "").border = border
        worksheet.cell(row=row_num, column=11, value=letter.due_date.strftime("%Y-%m-%d") if letter.due_date else "").border = border

    # Auto-adjust column widths
    column_widths = [18, 12, 12, 35, 25, 15, 12, 15, 18, 20, 12]
    for col_num, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    # Add title row at the top
    worksheet.insert_rows(1)
    worksheet.merge_cells('A1:K1')
    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14, color="4F46E5")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add generated date
    worksheet.insert_rows(2)
    worksheet.merge_cells('A2:K2')
    date_cell = worksheet.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    date_cell.font = Font(size=10, color="666666")
    date_cell.alignment = Alignment(horizontal="center", vertical="center")

    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{title.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    return response


def export_letters_to_pdf(letters, title="Letters Report"):
    """
    Export letters to PDF format using reportlab.
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{title.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d")}.pdf"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#4F46E5'),
        spaceAfter=20,
        alignment=1
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=30,
        alignment=1
    )

    elements = []
    
    # Title
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
    
    # Table data
    headers = ["Ref No", "Direction", "Date", "Subject", "Sender/Recipient", "Category", "Priority", "Status", "Dept", "Assigned To", "Due Date"]
    data = [headers]
    
    for letter in letters:
        row = [
            letter.reference_no or "",
            letter.get_direction_display(),
            letter.date.strftime("%Y-%m-%d") if letter.date else "",
            letter.subject[:50] + "..." if len(letter.subject) > 50 else letter.subject,
            (letter.sender if letter.direction == 'INCOMING' else letter.recipient) or "",
            letter.category.name if letter.category else "",
            letter.get_priority_display(),
            letter.get_status_display(),
            letter.assigned_department.name if letter.assigned_department else "",
            str(letter.assigned_person) if letter.assigned_person else "",
            letter.due_date.strftime("%Y-%m-%d") if letter.due_date else ""
        ]
        data.append(row)
    
    # Create table
    table = Table(data, colWidths=[0.8*inch, 0.8*inch, 0.8*inch, 1.5*inch, 1.2*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1.0*inch, 0.8*inch])
    
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')])
    ])
    
    table.setStyle(table_style)
    elements.append(table)
    
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    
    response.write(pdf)
    return response
