import calendar
import os

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count, Q
from django.http import JsonResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect
from django.utils import timezone
from django.views import View
from django.urls import reverse, reverse_lazy
from django.contrib.auth.models import User, Group
from django.views.generic import (
    CreateView, DetailView, ListView, TemplateView, UpdateView, DeleteView
)
from django_filters.views import FilterView
from django.db import models
 
from .filters import LetterFilter
from .forms import ActionLogForm, AttachmentForm, LetterForm, DepartmentForm, CategoryForm, StaffForm, IncomingLetterForm, OutgoingLetterForm, UserProfileForm, UserPreferencesForm, CustomPasswordChangeForm
from .models import ActionLog, Attachment, Department, Letter, Category, SavedSearch, UserProfile, Notification
from .permissions import (
    user_can_close, user_can_view_all_letters, CanViewLetterMixin, SuperuserOrAdminRequiredMixin,
)
from .export_utils import export_letters_to_excel, export_letters_to_pdf
from .email_utils import (
    send_overdue_notification, send_status_change_notification,
    send_assignment_notification, send_new_action_notification
)


# ---------------------------------------------------------------------------
# Notification Helper Functions
# ---------------------------------------------------------------------------
def create_notification(recipient, notification_type, title, message, related_letter=None):
    """Create a notification for a user."""
    # Check if user has email notifications enabled
    try:
        profile = recipient.profile
        if not profile.email_notifications:
            return
    except UserProfile.DoesNotExist:
        # Create profile if it doesn't exist
        profile = UserProfile.objects.create(user=recipient)
    
    Notification.objects.create(
        recipient=recipient,
        notification_type=notification_type,
        title=title,
        message=message,
        related_letter=related_letter
    )


# ---------------------------------------------------------------------------
# Letter Search API for autocomplete
# ---------------------------------------------------------------------------
class LetterSearchView(LoginRequiredMixin, View):
    """API endpoint for searching letters (used for related letter autocomplete)."""
    def get(self, request):
        query = request.GET.get('q', '')
        if not query or len(query) < 2:
            return JsonResponse({'results': []})
        
        letters_qs = get_user_letter_queryset(request.user)
        
        # Search by reference_no, sender, recipient, subject, department
        letters = letters_qs.filter(
            Q(reference_no__icontains=query) |
            Q(sender__icontains=query) |
            Q(recipient__icontains=query) |
            Q(subject__icontains=query) |
            Q(assigned_department__name__icontains=query)
        )[:10]  # Limit to 10 results
        
        results = []
        for letter in letters:
            results.append({
                'id': letter.pk,
                'text': f"{letter.reference_no} - {letter.subject[:50]}{'...' if len(letter.subject) > 50 else ''}",
                'reference_no': letter.reference_no,
                'subject': letter.subject,
                'sender': letter.sender or '',
                'recipient': letter.recipient or '',
            })
        
        return JsonResponse({'results': results})


def get_user_letter_queryset(user, qs=None):
    """Filters a letter queryset based on the user's role and assigned department."""
    if qs is None:
        qs = Letter.objects.all()
    if not user_can_view_all_letters(user):
        user_depts = user.departments.all()
        qs = qs.filter(
            Q(assigned_department__in=user_depts) |
            Q(assigned_person=user) |
            Q(created_by=user)
        )
    return qs


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------
class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = timezone.now().date()
        user = self.request.user

        letters_qs = get_user_letter_queryset(user)
        ctx['total_letters'] = letters_qs.count()
        ctx['incoming_count'] = letters_qs.filter(direction='INCOMING').count()
        ctx['outgoing_count'] = letters_qs.filter(direction='OUTGOING').count()
        ctx['pending_count'] = letters_qs.filter(
            status__in=['RECEIVED', 'DRAFTED', 'IN_REVIEW'],
        ).count()
        ctx['overdue_count'] = letters_qs.filter(
            due_date__lt=today,
        ).exclude(status__in=['ARCHIVED']).count()
        ctx['in_review_count'] = letters_qs.filter(status='IN_REVIEW').count()
        ctx['urgent_count'] = letters_qs.filter(
            priority='URGENT',
        ).exclude(status__in=['ARCHIVED']).count()

        # Filter action logs so user only sees actions for letters they can view
        action_qs = ActionLog.objects.all()
        if not user_can_view_all_letters(user):
            user_depts = user.departments.all()
            action_qs = action_qs.filter(
                Q(letter__assigned_department__in=user_depts) |
                Q(letter__assigned_person=user) |
                Q(letter__created_by=user)
            )

        ctx['recent_actions'] = (
            action_qs
            .select_related('letter', 'action_by')[:10]
        )
        ctx['recent_letters'] = (
            letters_qs
            .select_related('assigned_department', 'created_by')[:5]
        )
        return ctx


# ---------------------------------------------------------------------------
# Letter list (filterable)
# ---------------------------------------------------------------------------
class LetterListView(LoginRequiredMixin, FilterView):
    model = Letter
    template_name = 'letters/letter_list.html'
    context_object_name = 'letters'
    filterset_class = LetterFilter
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset()
        qs = get_user_letter_queryset(self.request.user, qs)
        qs = qs.select_related(
            'assigned_department', 'assigned_person', 'created_by',
        )
        search = self.request.GET.get('search', '')
        if search:
            qs = qs.filter(
                Q(subject__icontains=search)
                | Q(sender__icontains=search)
                | Q(recipient__icontains=search)
                | Q(reference_no__icontains=search)
            )
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['list_title'] = 'All Letters'
        ctx['list_clear_url'] = reverse_lazy('letters:letter_list')
        ctx['saved_searches'] = SavedSearch.objects.filter(user=self.request.user)[:10]
        return ctx

    def get(self, request, *args, **kwargs):
        # Handle export requests
        export_format = request.GET.get('export')
        if export_format in ('excel', 'pdf'):
            queryset = self.get_queryset()
            if export_format == 'excel':
                return export_letters_to_excel(queryset, 'All_Letters')
            elif export_format == 'pdf':
                return export_letters_to_pdf(queryset, 'All_Letters')
        return super().get(request, *args, **kwargs)


class OutgoingLetterListView(LetterListView):
    def get_queryset(self):
        return super().get_queryset().filter(direction='OUTGOING')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['list_title'] = 'Outgoing Letters'
        ctx['is_outgoing_view'] = True
        ctx['list_clear_url'] = reverse_lazy('letters:outgoing_letter_list')
        return ctx

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('export')
        if export_format in ('excel', 'pdf'):
            queryset = self.get_queryset()
            if export_format == 'excel':
                return export_letters_to_excel(queryset, 'Outgoing_Letters')
            elif export_format == 'pdf':
                return export_letters_to_pdf(queryset, 'Outgoing_Letters')
        return super().get(request, *args, **kwargs)


class IncomingLetterListView(LetterListView):
    def get_queryset(self):
        return super().get_queryset().filter(direction='INCOMING')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['list_title'] = 'Incoming Letters'
        ctx['is_incoming_view'] = True
        ctx['list_clear_url'] = reverse_lazy('letters:incoming_letter_list')
        return ctx

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('export')
        if export_format in ('excel', 'pdf'):
            queryset = self.get_queryset()
            if export_format == 'excel':
                return export_letters_to_excel(queryset, 'Incoming_Letters')
            elif export_format == 'pdf':
                return export_letters_to_pdf(queryset, 'Incoming_Letters')
        return super().get(request, *args, **kwargs)


# ---------------------------------------------------------------------------
# Letter detail
# ---------------------------------------------------------------------------
class LetterDetailView(LoginRequiredMixin, CanViewLetterMixin, DetailView):
    model = Letter
    template_name = 'letters/letter_detail.html'
    context_object_name = 'letter'

    def get_queryset(self):
        return super().get_queryset().select_related(
            'assigned_department', 'assigned_person',
            'created_by', 'related_letter', 'category',
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        letter = self.object
        can_close = user_can_close(self.request.user, letter)

        ctx['attachments'] = letter.attachments.select_related('uploaded_by')
        ctx['action_logs'] = letter.actions.select_related('action_by')
        ctx['action_form'] = ActionLogForm(can_close=can_close, letter=letter)
        ctx['attachment_form'] = AttachmentForm()
        ctx['can_close'] = can_close
        ctx['is_admin'] = self.request.user.is_superuser or self.request.user.groups.filter(name='Admin').exists()
        ctx['related_replies'] = letter.replies.select_related(
            'assigned_department', 'assigned_person',
        )
        ctx['action_count'] = ctx['action_logs'].count()
        
        # Set status flow based on letter direction
        if letter.direction == Letter.INCOMING:
            ctx['status_flow'] = [
                (val, label) for val, label in Letter.STATUS_CHOICES
                if val in ['RECEIVED', 'IN_REVIEW', 'ACTIONED', 'RESPONDED', 'CLOSED', 'ARCHIVED']
            ]
        elif letter.direction == Letter.OUTGOING:
            ctx['status_flow'] = [
                (val, label) for val, label in Letter.STATUS_CHOICES
                if val in ['DRAFTED', 'IN_REVIEW', 'SUBMITTED', 'RESPONDED', 'ARCHIVED']
            ]
        else:
            ctx['status_flow'] = Letter.STATUS_CHOICES

        if letter.direction == Letter.INCOMING:
            ctx['back_url'] = reverse('letters:incoming_letter_list')
            ctx['back_label'] = 'Incoming Letters'
        elif letter.direction == Letter.OUTGOING:
            ctx['back_url'] = reverse('letters:outgoing_letter_list')
            ctx['back_label'] = 'Outgoing Letters'
        else:
            ctx['back_url'] = reverse('letters:letter_list')
            ctx['back_label'] = 'All Letters'

        return ctx


# ---------------------------------------------------------------------------
# Create / Edit letter
# ---------------------------------------------------------------------------
class LetterCreateView(LoginRequiredMixin, CreateView):
    model = Letter
    form_class = LetterForm
    template_name = 'letters/letter_form.html'

    def get_form_kwargs(self):
        kw = super().get_form_kwargs()
        kw['user'] = self.request.user
        return kw

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        ActionLog.objects.create(
            letter=self.object,
            action=f'Letter created ({self.object.get_direction_display()})',
            action_by=self.request.user,
        )
        
        # Send assignment notification if assigned to someone
        if self.object.assigned_person:
            send_assignment_notification(self.object)
            # Create in-app notification
            create_notification(
                recipient=self.object.assigned_person,
                notification_type='letter_assigned',
                title=f'Letter Assigned: {self.object.reference_no}',
                message=f'You have been assigned to letter "{self.object.subject}"',
                related_letter=self.object
            )
        
        messages.success(
            self.request,
            f'Letter {self.object.reference_no} created successfully.',
        )
        return response

    def get_success_url(self):
        return self.object.get_absolute_url()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['form_title'] = 'Register New Letter'
        return ctx


class IncomingLetterCreateView(LetterCreateView):
    """Create view specifically for incoming letters."""
    form_class = IncomingLetterForm
    
    def get_initial(self):
        return {'direction': Letter.INCOMING}
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['form_title'] = 'Register Incoming Letter'
        return ctx
    
    def get_success_url(self):
        return reverse('letters:incoming_letter_list')


class OutgoingLetterCreateView(LetterCreateView):
    """Create view specifically for outgoing letters."""
    form_class = OutgoingLetterForm
    
    def get_initial(self):
        return {'direction': Letter.OUTGOING}
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['form_title'] = 'Register Outgoing Letter'
        return ctx
    
    def get_success_url(self):
        return reverse('letters:outgoing_letter_list')


class LetterUpdateView(LoginRequiredMixin, CanViewLetterMixin, UpdateView):
    model = Letter
    template_name = 'letters/letter_form.html'

    def get_form_class(self):
        """Use appropriate form based on letter direction."""
        if self.object.direction == Letter.INCOMING:
            return IncomingLetterForm
        elif self.object.direction == Letter.OUTGOING:
            return OutgoingLetterForm
        return LetterForm

    def get_form_kwargs(self):
        kw = super().get_form_kwargs()
        kw['user'] = self.request.user
        return kw

    def form_valid(self, form):
        old_assigned_person = self.object.assigned_person
        old_status = self.object.status
        
        response = super().form_valid(form)
        ActionLog.objects.create(
            letter=self.object,
            action='Letter details updated',
            action_by=self.request.user,
        )
        
        # Send assignment notification if assigned person changed
        if self.object.assigned_person and self.object.assigned_person != old_assigned_person:
            send_assignment_notification(self.object)
            # Create in-app notification
            create_notification(
                recipient=self.object.assigned_person,
                notification_type='letter_assigned',
                title=f'Letter Assigned: {self.object.reference_no}',
                message=f'You have been assigned to letter "{self.object.subject}"',
                related_letter=self.object
            )
        
        # Send status change notification if status changed
        if self.object.status != old_status:
            send_status_change_notification(self.object, old_status, self.object.status)
            # Create in-app notification for assigned person
            if self.object.assigned_person:
                create_notification(
                    recipient=self.object.assigned_person,
                    notification_type='status_changed',
                    title=f'Status Changed: {self.object.reference_no}',
                    message=f'Letter status changed from {old_status} to {self.object.status}',
                    related_letter=self.object
                )
        
        messages.success(
            self.request,
            f'Letter {self.object.reference_no} updated.',
        )
        return response

    def get_success_url(self):
        return self.object.get_absolute_url()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['form_title'] = f'Edit: {self.object.reference_no}'
        return ctx


class LetterDeleteView(LoginRequiredMixin, DeleteView):
    """Delete a letter - only for superusers."""
    model = Letter
    template_name = 'letters/letter_confirm_delete.html'
    success_url = reverse_lazy('letters:letter_list')

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            messages.error(request, 'Only superusers can delete letters.')
            return redirect('letters:letter_detail', pk=self.get_object().pk)
        return super().dispatch(request, *args, **kwargs)

    def delete(self, request, *args, **kwargs):
        letter = self.get_object()
        messages.success(request, f'Letter {letter.reference_no} has been deleted.')
        return super().delete(request, *args, **kwargs)


# ---------------------------------------------------------------------------
# Action log (POST from detail page)
# ---------------------------------------------------------------------------
class AddActionView(LoginRequiredMixin, View):
    def post(self, request, pk):
        letters_qs = get_user_letter_queryset(request.user)
        letter = get_object_or_404(letters_qs, pk=pk)
        form = ActionLogForm(request.POST)

        if form.is_valid():
            action_log = form.save(commit=False)
            action_log.letter = letter
            action_log.action_by = request.user
            action_log.save()

            # Send notification for new action
            send_new_action_notification(letter, action_log.action, request.user)
            # Create in-app notification for assigned person
            if letter.assigned_person and letter.assigned_person != request.user:
                create_notification(
                    recipient=letter.assigned_person,
                    notification_type='action_added',
                    title=f'Action Added: {letter.reference_no}',
                    message=f'New action added: "{action_log.action}"',
                    related_letter=letter
                )

            # Optionally update letter status
            new_status = form.cleaned_data.get('new_status')
            if new_status:
                if new_status in ('CLOSED', 'ARCHIVED'):
                    if not user_can_close(request.user, letter):
                        messages.error(
                            request,
                            'Only the assigned person or admin can '
                            'close/archive letters.',
                        )
                        return redirect(letter.get_absolute_url())

                old_display = letter.get_status_display()
                letter.status = new_status
                letter.save(update_fields=['status', 'updated_at'])

                new_display = letter.get_status_display()
                ActionLog.objects.create(
                    letter=letter,
                    action=f'Status changed: {old_display} → {new_display}',
                    action_by=request.user,
                )
                
                # Send status change notification
                send_status_change_notification(letter, old_display, new_display)

            messages.success(request, 'Action logged successfully.')
        else:
            messages.error(request, 'Error logging action. Please check the form.')

        return redirect(letter.get_absolute_url())


# ---------------------------------------------------------------------------
# Attachment upload (POST from detail page)
# ---------------------------------------------------------------------------
class AddAttachmentView(LoginRequiredMixin, View):
    def post(self, request, pk):
        letters_qs = get_user_letter_queryset(request.user)
        letter = get_object_or_404(letters_qs, pk=pk)
        form = AttachmentForm(request.POST, request.FILES)

        if form.is_valid():
            try:
                att = form.save(commit=False)
                att.letter = letter
                att.uploaded_by = request.user
                
                # Debug logging
                print(f"Storage backend: {att.file.storage}")
                print(f"File name before save: {att.file.name}")
                print(f"USE_R2_STORAGE: {os.environ.get('USE_R2_STORAGE')}")
                
                att.save()
                
                print(f"File name after save: {att.file.name}")
                print(f"File URL: {att.file.url}")

                ActionLog.objects.create(
                    letter=letter,
                    action=f'Attachment added: {att.filename}',
                    action_by=request.user,
                )
                messages.success(request, 'Attachment uploaded successfully.')
            except Exception as e:
                print(f"Error saving attachment: {e}")
                import traceback
                traceback.print_exc()
                messages.error(request, f'Error saving attachment: {str(e)}')
        else:
            # Display specific form validation errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')

        return redirect(letter.get_absolute_url())


# ---------------------------------------------------------------------------
# Attachment delete (for admin users)
# ---------------------------------------------------------------------------
class AttachmentDeleteView(LoginRequiredMixin, View):
    """Delete an attachment - only for admin/superuser users."""
    def post(self, request, letter_pk, attachment_pk):
        letters_qs = get_user_letter_queryset(request.user)
        letter = get_object_or_404(letters_qs, pk=letter_pk)
        
        # Check if user is admin or superuser
        if not (request.user.is_superuser or request.user.groups.filter(name='Admin').exists()):
            messages.error(request, 'Only admin users can delete attachments.')
            return redirect(letter.get_absolute_url())
        
        attachment = get_object_or_404(letter.attachments, pk=attachment_pk)
        filename = attachment.filename
        attachment.delete()
        
        ActionLog.objects.create(
            letter=letter,
            action=f'Attachment deleted: {filename}',
            action_by=request.user,
        )
        messages.success(request, f'Attachment "{filename}" deleted successfully.')
        
        return redirect(letter.get_absolute_url())


# ---------------------------------------------------------------------------
# Overdue letters
# ---------------------------------------------------------------------------
class OverdueLettersView(LoginRequiredMixin, ListView):
    template_name = 'letters/overdue_list.html'
    context_object_name = 'letters'
    paginate_by = 20

    def get_queryset(self):
        today = timezone.now().date()
        qs = Letter.objects.filter(due_date__lt=today).exclude(status__in=['ARCHIVED'])
        qs = get_user_letter_queryset(self.request.user, qs)
        return qs.select_related('assigned_department', 'assigned_person')


# ---------------------------------------------------------------------------
# Bulk Actions
# ---------------------------------------------------------------------------
class BulkActionView(LoginRequiredMixin, View):
    """Handle bulk actions on multiple letters."""
    
    def post(self, request):
        action = request.POST.get('action')
        letter_ids = request.POST.getlist('letter_ids')
        
        if not letter_ids:
            messages.error(request, 'No letters selected.')
            return redirect(request.META.get('HTTP_REFERER', reverse('letters:letter_list')))
        
        letters_qs = get_user_letter_queryset(request.user)
        letters = Letter.objects.filter(pk__in=letter_ids)
        
        # Ensure user can only act on letters they have access to
        accessible_ids = set(letters_qs.filter(pk__in=letter_ids).values_list('pk', flat=True))
        letters = letters.filter(pk__in=accessible_ids)
        
        count = 0
        
        if action == 'delete':
            if not request.user.is_superuser:
                messages.error(request, 'Only superusers can delete letters.')
                return redirect(request.META.get('HTTP_REFERER', reverse('letters:letter_list')))
            
            count = letters.count()
            letters.delete()
            messages.success(request, f'{count} letter(s) deleted successfully.')
        
        elif action == 'close':
            for letter in letters:
                if user_can_close(request.user, letter):
                    old_status = letter.status
                    # Use CLOSED for incoming, ARCHIVED for outgoing
                    if letter.direction == Letter.INCOMING:
                        new_status = 'CLOSED'
                    else:
                        new_status = 'ARCHIVED'
                    letter.status = new_status
                    letter.save(update_fields=['status', 'updated_at'])
                    ActionLog.objects.create(
                        letter=letter,
                        action=f'Status changed: {old_status} → {new_status} (bulk action)',
                        action_by=request.user,
                    )
                    send_status_change_notification(letter, old_status, new_status)
                    count += 1
            messages.success(request, f'{count} letter(s) closed/archived successfully.')
        
        elif action == 'archive':
            for letter in letters:
                if user_can_close(request.user, letter):
                    old_status = letter.status
                    letter.status = 'ARCHIVED'
                    letter.save(update_fields=['status', 'updated_at'])
                    ActionLog.objects.create(
                        letter=letter,
                        action=f'Status changed: {old_status} → ARCHIVED (bulk action)',
                        action_by=request.user,
                    )
                    send_status_change_notification(letter, old_status, 'ARCHIVED')
                    count += 1
            messages.success(request, f'{count} letter(s) archived successfully.')
        
        elif action == 'reopen':
            for letter in letters:
                old_status = letter.status
                letter.status = 'IN_REVIEW'
                letter.save(update_fields=['status', 'updated_at'])
                ActionLog.objects.create(
                    letter=letter,
                    action=f'Status changed: {old_status} → IN_REVIEW (bulk action)',
                    action_by=request.user,
                )
                send_status_change_notification(letter, old_status, 'IN_REVIEW')
                count += 1
            messages.success(request, f'{count} letter(s) reopened successfully.')
        
        elif action.startswith('assign_to_'):
            user_id = action.replace('assign_to_', '')
            try:
                assign_to = User.objects.get(pk=user_id)
                for letter in letters:
                    old_assigned = letter.assigned_person
                    letter.assigned_person = assign_to
                    letter.save(update_fields=['assigned_person', 'updated_at'])
                    ActionLog.objects.create(
                        letter=letter,
                        action=f'Assigned to {assign_to.get_full_name() or assign_to.username} (bulk action)',
                        action_by=request.user,
                    )
                    if old_assigned != assign_to:
                        send_assignment_notification(letter)
                    count += 1
                messages.success(request, f'{count} letter(s) assigned to {assign_to.get_full_name() or assign_to.username}.')
            except User.DoesNotExist:
                messages.error(request, 'Invalid user selected.')
        
        else:
            messages.error(request, 'Invalid action selected.')
        
        return redirect(request.META.get('HTTP_REFERER', reverse('letters:letter_list')))


# ---------------------------------------------------------------------------
# Saved Searches
# ---------------------------------------------------------------------------
class SaveSearchView(LoginRequiredMixin, View):
    """Save current search query for quick access."""
    
    def post(self, request):
        name = request.POST.get('name')
        query_string = request.GET.urlencode()
        
        if not name:
            messages.error(request, 'Please provide a name for your saved search.')
            return redirect(request.META.get('HTTP_REFERER', reverse('letters:letter_list')))
        
        # Remove empty parameters
        params = request.GET.copy()
        for key in list(params.keys()):
            if not params[key]:
                del params[key]
        query_string = params.urlencode()
        
        if not query_string:
            messages.error(request, 'Cannot save an empty search.')
            return redirect(request.META.get('HTTP_REFERER', reverse('letters:letter_list')))
        
        SavedSearch.objects.update_or_create(
            user=request.user,
            name=name,
            defaults={'query_string': query_string}
        )
        
        messages.success(request, f'Search "{name}" saved successfully.')
        return redirect(request.META.get('HTTP_REFERER', reverse('letters:letter_list')))


class DeleteSearchView(LoginRequiredMixin, DeleteView):
    """Delete a saved search."""
    model = SavedSearch
    template_name = 'includes/confirm_delete.html'
    
    def get_queryset(self):
        return SavedSearch.objects.filter(user=self.request.user)
    
    def get_success_url(self):
        return reverse('letters:letter_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Saved search deleted successfully.')
        return super().delete(request, *args, **kwargs)


class ExportActionsView(LoginRequiredMixin, View):
    """Export action logs for a letter to Excel/PDF."""
    
    def get(self, request, pk):
        letters_qs = get_user_letter_queryset(request.user)
        letter = get_object_or_404(letters_qs, pk=pk)
        actions = letter.actions.select_related('action_by')
        
        export_format = request.GET.get('export', 'excel')
        
        if export_format == 'excel':
            from .export_utils import export_letters_to_excel
            # Create a simple Excel export for actions
            from openpyxl import Workbook
            from io import BytesIO
            from django.http import HttpResponse
            
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Actions"
            
            # Headers
            headers = ['Date', 'Action', 'Action By', 'Notes']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            
            # Data
            for row_num, action in enumerate(actions, 2):
                ws.cell(row=row_num, column=1, value=action.action_date.strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row_num, column=2, value=action.action)
                ws.cell(row=row_num, column=3, value=str(action.action_by) if action.action_by else '')
                ws.cell(row=row_num, column=4, value=action.notes or '')
            
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="actions_{letter.reference_no}.xlsx"'
            return response
        
        return redirect(letter.get_absolute_url())


# ---------------------------------------------------------------------------
# Reports
# ---------------------------------------------------------------------------
class ReportsView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/reports.html'

    def get(self, request, *args, **kwargs):
        # Handle export requests
        export_format = request.GET.get('export')
        if export_format in ('excel', 'pdf'):
            from .export_utils import export_letters_to_excel, export_letters_to_pdf
            from .models import Letter
            
            base_qs = get_user_letter_queryset(request.user, Letter.objects.all())
            
            date_from = request.GET.get('date_from')
            date_to = request.GET.get('date_to')
            
            if date_from:
                base_qs = base_qs.filter(date__gte=date_from)
            if date_to:
                base_qs = base_qs.filter(date__lte=date_to)
            
            if export_format == 'excel':
                return export_letters_to_excel(base_qs, 'Reports_Data')
            elif export_format == 'pdf':
                return export_letters_to_pdf(base_qs, 'Reports_Data')
        return super().get(request, *args, **kwargs)


class ReportsDataView(LoginRequiredMixin, View):
    """JSON endpoint consumed by Chart.js on the reports page."""

    def get(self, request):
        today = timezone.now().date()
        user = request.user
        base_qs = get_user_letter_queryset(user)
        
        # Apply date range filters
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        
        if date_from:
            base_qs = base_qs.filter(date__gte=date_from)
        if date_to:
            base_qs = base_qs.filter(date__lte=date_to)

        # -- Monthly volume (last 12 months) ---------------------------------
        monthly = []
        for i in range(11, -1, -1):
            m = today.month - i
            y = today.year
            while m <= 0:
                m += 12
                y -= 1
            incoming = base_qs.filter(
                direction='INCOMING', date__year=y, date__month=m,
            ).count()
            outgoing = base_qs.filter(
                direction='OUTGOING', date__year=y, date__month=m,
            ).count()
            monthly.append({
                'month': f'{calendar.month_abbr[m]} {y}',
                'incoming': incoming,
                'outgoing': outgoing,
            })

        # -- By department ---------------------------------------------------
        dept_data = list(
            base_qs
            .values('assigned_department__name')
            .annotate(count=Count('id'))
            .order_by('-count')
        )

        # -- By category -----------------------------------------------------
        cat_data = [
            {'category': item['category__name'] or 'Uncategorized', 'count': item['count']}
            for item in base_qs.values('category__name').annotate(count=Count('id')).order_by('-count')
        ]

        # -- By status -------------------------------------------------------
        status_data = list(
            base_qs
            .values('status')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        
        # -- By priority -----------------------------------------------------
        priority_data = list(
            base_qs
            .values('priority')
            .annotate(count=Count('id'))
            .order_by('-count')
        )

        return JsonResponse({
            'monthly': monthly,
            'by_department': dept_data,
            'by_category': cat_data,
            'by_status': status_data,
            'by_priority': priority_data,
        })


# ---------------------------------------------------------------------------
# Superadmin / System Admin Settings Dashboard and CRUD views
# ---------------------------------------------------------------------------

class AdminDashboardView(LoginRequiredMixin, SuperuserOrAdminRequiredMixin, TemplateView):
    template_name = 'admin_dashboard/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['total_departments'] = Department.objects.count()
        ctx['total_categories'] = Category.objects.count()
        ctx['total_staff'] = User.objects.filter(is_active=True).count()
        ctx['total_letters'] = Letter.objects.count()
        ctx['recent_staff'] = User.objects.order_by('-date_joined')[:5]
        ctx['recent_letters'] = Letter.objects.order_by('-created_at')[:5]
        return ctx


# --- Department CRUD ---
class DepartmentListView(LoginRequiredMixin, SuperuserOrAdminRequiredMixin, ListView):
    model = Department
    template_name = 'admin_dashboard/department_list.html'
    context_object_name = 'departments'
    paginate_by = 10

    def get_queryset(self):
        return Department.objects.annotate(staff_count=Count('users')).order_by('name')


class DepartmentCreateView(LoginRequiredMixin, SuperuserOrAdminRequiredMixin, CreateView):
    model = Department
    form_class = DepartmentForm
    template_name = 'admin_dashboard/department_form.html'
    success_url = reverse_lazy('letters:admin_department_list')

    def form_valid(self, form):
        messages.success(self.request, f"Department '{form.instance.name}' created successfully.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Create Department'
        return ctx


class DepartmentUpdateView(LoginRequiredMixin, SuperuserOrAdminRequiredMixin, UpdateView):
    model = Department
    form_class = DepartmentForm
    template_name = 'admin_dashboard/department_form.html'
    success_url = reverse_lazy('letters:admin_department_list')

    def form_valid(self, form):
        messages.success(self.request, f"Department '{form.instance.name}' updated.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = f"Edit Department: {self.object.name}"
        return ctx


class DepartmentDeleteView(LoginRequiredMixin, SuperuserOrAdminRequiredMixin, DeleteView):
    model = Department
    template_name = 'admin_dashboard/department_confirm_delete.html'
    success_url = reverse_lazy('letters:admin_department_list')

    def form_valid(self, form):
        messages.success(self.request, f"Department '{self.object.name}' deleted.")
        return super().form_valid(form)


# --- Category CRUD ---
class CategoryListView(LoginRequiredMixin, SuperuserOrAdminRequiredMixin, ListView):
    model = Category
    template_name = 'admin_dashboard/category_list.html'
    context_object_name = 'categories'
    paginate_by = 10

    def get_queryset(self):
        return Category.objects.annotate(letter_count=Count('letters')).order_by('name')


class CategoryCreateView(LoginRequiredMixin, SuperuserOrAdminRequiredMixin, CreateView):
    model = Category
    form_class = CategoryForm
    template_name = 'admin_dashboard/category_form.html'
    success_url = reverse_lazy('letters:admin_category_list')

    def form_valid(self, form):
        messages.success(self.request, f"Category '{form.instance.name}' created successfully.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Create Category'
        return ctx


class CategoryUpdateView(LoginRequiredMixin, SuperuserOrAdminRequiredMixin, UpdateView):
    model = Category
    form_class = CategoryForm
    template_name = 'admin_dashboard/category_form.html'
    success_url = reverse_lazy('letters:admin_category_list')

    def form_valid(self, form):
        messages.success(self.request, f"Category '{form.instance.name}' updated.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = f"Edit Category: {self.object.name}"
        return ctx


class CategoryDeleteView(LoginRequiredMixin, SuperuserOrAdminRequiredMixin, DeleteView):
    model = Category
    template_name = 'admin_dashboard/category_confirm_delete.html'
    success_url = reverse_lazy('letters:admin_category_list')

    def form_valid(self, form):
        messages.success(self.request, f"Category '{self.object.name}' deleted.")
        return super().form_valid(form)


# --- Staff CRUD ---
class StaffListView(LoginRequiredMixin, SuperuserOrAdminRequiredMixin, ListView):
    model = User
    template_name = 'admin_dashboard/staff_list.html'
    context_object_name = 'staff_members'
    paginate_by = 10

    def get_queryset(self):
        return User.objects.prefetch_related('departments', 'groups').order_by('-is_active', 'username')


class StaffCreateView(LoginRequiredMixin, SuperuserOrAdminRequiredMixin, CreateView):
    model = User
    form_class = StaffForm
    template_name = 'admin_dashboard/staff_form.html'
    success_url = reverse_lazy('letters:admin_staff_list')

    def form_valid(self, form):
        messages.success(self.request, f"Staff account '{form.instance.username}' registered successfully.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Register Staff Member'
        return ctx


class StaffUpdateView(LoginRequiredMixin, SuperuserOrAdminRequiredMixin, UpdateView):
    model = User
    form_class = StaffForm
    template_name = 'admin_dashboard/staff_form.html'
    success_url = reverse_lazy('letters:admin_staff_list')

    def form_valid(self, form):
        messages.success(self.request, f"Staff account '{form.instance.username}' updated.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = f"Edit Staff Account: {self.object.username}"
        return ctx


class StaffDeleteView(LoginRequiredMixin, SuperuserOrAdminRequiredMixin, DeleteView):
    model = User
    template_name = 'admin_dashboard/staff_confirm_delete.html'
    success_url = reverse_lazy('letters:admin_staff_list')

    def form_valid(self, form):
        messages.success(self.request, f"Staff account '{self.object.username}' deleted.")
        return super().form_valid(form)


# ---------------------------------------------------------------------------
# Short URL redirect for file sharing
# ---------------------------------------------------------------------------
class ShortUrlRedirectView(View):
    """Redirect short URL to actual file URL with analytics."""
    def get(self, request, short_code):
        # Try to find by custom short code first, then regular short code
        attachment = get_object_or_404(Attachment, models.Q(short_code=short_code) | models.Q(custom_short_code=short_code))
        
        # Check if link has expired
        if attachment.is_expired():
            from django.http import HttpResponse
            return HttpResponse("This link has expired.", status=410)
        
        # Update analytics
        attachment.access_count += 1
        attachment.last_accessed = timezone.now()
        attachment.save(update_fields=['access_count', 'last_accessed'])
        
        return HttpResponseRedirect(attachment.file.url)


# ---------------------------------------------------------------------------
# User Profile Management
# ---------------------------------------------------------------------------
class UserProfileView(LoginRequiredMixin, TemplateView):
    """User profile page with profile info, preferences, and password change."""
    template_name = 'letters/user_profile.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Get or create user profile
        profile, created = UserProfile.objects.get_or_create(user=user)
        
        context['profile'] = profile
        context['profile_form'] = UserProfileForm(instance=user)
        context['preferences_form'] = UserPreferencesForm(instance=profile)
        context['password_form'] = CustomPasswordChangeForm(user=user)
        
        # Calculate profile completion
        completion = 0
        if user.first_name:
            completion += 25
        if user.last_name:
            completion += 25
        if user.email:
            completion += 25
        if profile.avatar:
            completion += 25
        context['profile_completion'] = completion
        
        # Calculate activity stats
        from .models import Letter, Attachment
        
        letters_assigned = Letter.objects.filter(assigned_person=user).count()
        letters_completed = Letter.objects.filter(
            assigned_person=user,
            status__in=['RESPONDED', 'CLOSED', 'ARCHIVED']
        ).count()
        letters_pending = letters_assigned - letters_completed
        attachments_count = Attachment.objects.filter(uploaded_by=user).count()
        
        context['letters_assigned'] = letters_assigned
        context['letters_completed'] = letters_completed
        context['letters_pending'] = letters_pending
        context['attachments_count'] = attachments_count
        
        return context
    
    def post(self, request, *args, **kwargs):
        user = request.user
        profile, created = UserProfile.objects.get_or_create(user=user)
        
        form_type = request.POST.get('form_type')
        
        if form_type == 'profile':
            form = UserProfileForm(request.POST, instance=user)
            if form.is_valid():
                form.save()
                messages.success(request, 'Profile updated successfully.')
                return redirect('letters:user_profile')
        
        elif form_type == 'preferences':
            form = UserPreferencesForm(request.POST, request.FILES, instance=profile)
            if form.is_valid():
                form.save()
                messages.success(request, 'Preferences saved successfully.')
                return redirect('letters:user_profile')
        
        elif form_type == 'password':
            form = CustomPasswordChangeForm(user=user, data=request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'Password changed successfully.')
                return redirect('letters:user_profile')
        
        # If form is invalid, re-render with errors
        context = self.get_context_data(**kwargs)
        if form_type == 'profile':
            context['profile_form'] = form
        elif form_type == 'preferences':
            context['preferences_form'] = form
        elif form_type == 'password':
            context['password_form'] = form
        
        return self.render_to_response(context)


# ---------------------------------------------------------------------------
# Notification Management
# ---------------------------------------------------------------------------
class NotificationListView(LoginRequiredMixin, ListView):
    """List of user notifications."""
    model = Notification
    template_name = 'letters/notification_list.html'
    context_object_name = 'notifications'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Notification.objects.filter(
            recipient=self.request.user
        ).select_related('related_letter')
        
        # Filter by notification type
        notification_type = self.request.GET.get('type')
        if notification_type:
            queryset = queryset.filter(notification_type=notification_type)
        
        # Filter by read status
        read_status = self.request.GET.get('read')
        if read_status == 'unread':
            queryset = queryset.filter(is_read=False)
        elif read_status == 'read':
            queryset = queryset.filter(is_read=True)
        
        # Check if grouping is enabled
        self.group_by_letter = self.request.GET.get('group') == 'letter'
        
        if self.group_by_letter:
            # Group by related letter
            from django.db.models import Max
            grouped = {}
            for notification in queryset:
                letter_id = notification.related_letter.id if notification.related_letter else None
                if letter_id not in grouped:
                    grouped[letter_id] = []
                grouped[letter_id].append(notification)
            
            # Sort groups by most recent notification
            sorted_groups = sorted(
                grouped.items(),
                key=lambda x: max(n.created_at for n in x[1]),
                reverse=True
            )
            
            # Flatten back to queryset-like structure
            grouped_notifications = []
            for letter_id, notifications in sorted_groups:
                # Sort notifications within group by created_at
                notifications = sorted(notifications, key=lambda x: x.created_at, reverse=True)
                grouped_notifications.extend(notifications)
            
            # Store grouped data for template
            self.grouped_data = grouped
            return grouped_notifications
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['unread_count'] = Notification.objects.filter(
            recipient=self.request.user, is_read=False
        ).count()
        
        # Add filter context
        context['current_type'] = self.request.GET.get('type', '')
        context['current_read'] = self.request.GET.get('read', '')
        context['current_group'] = self.request.GET.get('group', '')
        context['notification_types'] = Notification.NOTIFICATION_TYPES
        context['group_by_letter'] = getattr(self, 'group_by_letter', False)
        context['grouped_data'] = getattr(self, 'grouped_data', {})
        
        return context


class NotificationDetailView(LoginRequiredMixin, DetailView):
    """View single notification and mark as read."""
    model = Notification
    template_name = 'letters/notification_detail.html'
    
    def get_queryset(self):
        return Notification.objects.filter(recipient=self.request.user)
    
    def get(self, request, *args, **kwargs):
        response = super().get(request, *args, **kwargs)
        notification = self.get_object()
        notification.mark_as_read()
        return response


class MarkAsReadView(LoginRequiredMixin, View):
    """Mark notification as read via AJAX."""
    def post(self, request, pk):
        notification = get_object_or_404(Notification, pk=pk, recipient=request.user)
        notification.mark_as_read()
        return JsonResponse({'success': True})


class MarkAllAsReadView(LoginRequiredMixin, View):
    """Mark all notifications as read."""
    def post(self, request):
        Notification.objects.filter(
            recipient=request.user, is_read=False
        ).update(is_read=True, read_at=timezone.now())
        return JsonResponse({'success': True})


class NotificationDeleteView(LoginRequiredMixin, View):
    """Delete a notification."""
    def post(self, request, pk):
        notification = get_object_or_404(Notification, pk=pk, recipient=request.user)
        notification.delete()
        return JsonResponse({'success': True})


class NotificationAPIView(LoginRequiredMixin, View):
    """API endpoint for loading notifications via AJAX."""
    def get(self, request):
        notifications = Notification.objects.filter(
            recipient=request.user
        ).select_related('related_letter')[:10]
        
        notification_data = []
        for notification in notifications:
            # Calculate time ago
            from datetime import timedelta
            time_diff = timezone.now() - notification.created_at
            
            if time_diff < timedelta(minutes=1):
                time_ago = 'Just now'
            elif time_diff < timedelta(hours=1):
                time_ago = f'{int(time_diff.total_seconds() / 60)} min ago'
            elif time_diff < timedelta(days=1):
                time_ago = f'{int(time_diff.total_seconds() / 3600)} hours ago'
            else:
                time_ago = f'{time_diff.days} days ago'
            
            # Get icon based on notification type
            icon_map = {
                'letter_assigned': 'bi-envelope',
                'status_changed': 'bi-arrow-repeat',
                'action_added': 'bi-list-check',
                'attachment_added': 'bi-paperclip',
                'overdue_warning': 'bi-exclamation-triangle',
                'comment_added': 'bi-chat',
            }
            
            color_map = {
                'letter_assigned': 'primary',
                'status_changed': 'info',
                'action_added': 'success',
                'attachment_added': 'warning',
                'overdue_warning': 'danger',
                'comment_added': 'secondary',
            }
            
            notification_data.append({
                'id': notification.id,
                'title': notification.title,
                'message': notification.message,
                'is_read': notification.is_read,
                'time_ago': time_ago,
                'icon': icon_map.get(notification.notification_type, 'bi-bell'),
                'color': color_map.get(notification.notification_type, 'primary'),
                'url': notification.related_letter.get_absolute_url() if notification.related_letter else '#',
            })
        
        unread_count = Notification.objects.filter(
            recipient=request.user, is_read=False
        ).count()
        
        return JsonResponse({
            'notifications': notification_data,
            'unread_count': unread_count,
        })

